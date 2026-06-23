"""
title: Spreadsheet Creator
author: open-webui-aws-fargate
version: 2.0.0
description: Generate a Microsoft Excel (.xlsx) workbook from one or more sheets, each supplied as a CSV string. The file is delivered as a chat attachment.
required_open_webui_version: 0.9.0
"""

import csv as _csv
import io
import re
import uuid
from typing import Optional

from pydantic import BaseModel, Field, ValidationError

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class _SheetInput(BaseModel):
    name: str = Field(description="Sheet tab name (max 31 characters).")
    csv: str = Field(description="Sheet data as a CSV string. The first row is treated as column headers and rendered bold.")


class Tools:
    class Valves(BaseModel):
        max_sheets: int = Field(default=20, description="Refuse workbooks with more than this many sheets.")
        max_rows_per_sheet: int = Field(default=10000, description="Refuse sheets with more than this many data rows.")
        max_size_mb: int = Field(default=15, description="Refuse to attach files larger than this many megabytes.")

    def __init__(self):
        self.valves = self.Valves()
        self.citation = False

    async def create_spreadsheet(
        self,
        sheets: list[dict],
        filename: str = "spreadsheet",
        __event_emitter__=None,
        __user__: Optional[dict] = None,
    ) -> str:
        """
        Create a Microsoft Excel (.xlsx) workbook and attach it to the chat.

        Args:
            sheets: List of sheet objects, each with:
                - name (string, required): tab name shown at the bottom of the workbook
                - csv (string, required): sheet data as CSV text. First row becomes bold
                  column headers. Use standard comma-separated values; quote fields that
                  contain commas or newlines.
            filename: Base name for the output file (without extension). Defaults to "spreadsheet".
        """
        if not isinstance(sheets, list) or not sheets:
            return "Error: 'sheets' must be a non-empty list of sheet objects."
        if len(sheets) > self.valves.max_sheets:
            return f"Error: too many sheets ({len(sheets)}); limit is {self.valves.max_sheets}."

        try:
            validated = [_SheetInput(**s) for s in sheets]
        except (ValidationError, TypeError) as exc:
            return f"Error parsing sheets: {exc}"

        wb = Workbook()
        wb.remove(wb.active)

        for sheet_def in validated:
            rows = _parse_csv(sheet_def.csv)
            if len(rows) - 1 > self.valves.max_rows_per_sheet:
                return (
                    f"Error: sheet '{sheet_def.name}' has {len(rows) - 1} data rows; "
                    f"limit is {self.valves.max_rows_per_sheet}."
                )

            ws = wb.create_sheet(title=(sheet_def.name or "Sheet")[:31])

            for row_idx, row in enumerate(rows, start=1):
                ws.append(row)
                if row_idx == 1:
                    for cell in ws[1]:
                        cell.font = Font(bold=True)

            # Auto-size columns based on content (cap at 60 chars wide).
            if rows:
                headers = rows[0]
                for col_idx, header in enumerate(headers, start=1):
                    width = max(len(str(header)), 8)
                    for data_row in rows[1:201]:
                        if col_idx - 1 < len(data_row) and data_row[col_idx - 1] is not None:
                            width = max(width, min(len(str(data_row[col_idx - 1])), 60))
                    ws.column_dimensions[get_column_letter(col_idx)].width = width + 2

        buffer = io.BytesIO()
        wb.save(buffer)
        data = buffer.getvalue()

        return await _attach_file(
            __event_emitter__,
            __user__,
            data=data,
            filename=f"{_safe_filename(filename)}.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            max_size_mb=self.valves.max_size_mb,
        )


def _parse_csv(text: str) -> list[list]:
    """Parse a CSV string into a list of rows (list of lists)."""
    reader = _csv.reader(io.StringIO(text.strip()))
    return [row for row in reader if any(cell.strip() for cell in row)]


def _safe_filename(name: str, fallback: str = "spreadsheet") -> str:
    name = re.sub(r"[^\w\-. ]", "_", name).strip()
    return name[:120] or fallback


async def _attach_file(
    event_emitter,
    user: Optional[dict],
    *,
    data: bytes,
    filename: str,
    content_type: str,
    max_size_mb: int = 10,
) -> str:
    if event_emitter is None:
        return f"(no event emitter; would attach {filename}, {len(data)} bytes)"
    if len(data) > max_size_mb * 1024 * 1024:
        return f"Refusing to attach: {filename} is {len(data) / 1e6:.1f} MB (limit {max_size_mb} MB)."

    from open_webui.models.files import Files, FileForm
    from open_webui.storage.provider import Storage

    file_id = str(uuid.uuid4())
    user_id = (user or {}).get("id", "")
    storage_filename = f"{file_id}_{filename}"

    _, storage_path = Storage.upload_file(io.BytesIO(data), storage_filename, {})

    record = await Files.insert_new_file(
        user_id,
        FileForm(
            id=file_id,
            filename=filename,
            path=storage_path,
            meta={"name": filename, "content_type": content_type, "size": len(data)},
        ),
    )
    if record is None:
        return f"Failed to register file {filename} with Open WebUI."

    await event_emitter(
        {
            "type": "files",
            "data": {
                "files": [
                    {
                        "type": "file",
                        "id": record.id,
                        "url": f"/api/v1/files/{record.id}",
                        "name": filename,
                        "size": len(data),
                        "status": "uploaded",
                        "error": "",
                        "itemId": str(uuid.uuid4()),
                    }
                ],
            },
        }
    )

    return f"Created **{filename}** ({len(data):,} bytes). The file is now attached to the chat."
